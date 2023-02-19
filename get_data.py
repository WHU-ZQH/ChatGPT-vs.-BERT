import os
import openpyxl as op
import sys
from random import sample
import argparse

task2data={
    "cola":"glue_data/CoLA/dev.tsv",
    "sst2":"glue_data/SST-2/dev.tsv",
    "mrpc":"glue_data/MRPC/dev.tsv",
    "stsb":"glue_data/STS-B/dev.tsv",
    "qqp":"glue_data/QQP/dev.tsv",
    "mnli":["glue_data/MNLI/dev_matched.tsv","glue_data/MNLI/dev_mismatched.tsv"],
    "qnli":"glue_data/QNLI/dev.tsv",
    "rte":"glue_data/RTE/dev.tsv",
    "wnli":"Superglue_data/WSC/dev.tsv",
}
task2prompt={
    "cola":'For the sentence: "[text]", is the sentence grammarly correct?',
    "sst2":'For the sentence: "[text]", is the sentiment in this sentence positive or negative?',
    "mrpc":'For the sentence pair "[text1]" and "[text2]", do these two sentences have the same semantics?',
    "stsb":'Determine the similarity between the following two sentences: "[text1]" and "[text2]". The score should be ranging from 0.0 to 5.0, and can be a decimal."',
    "qqp":'For the sentence pair "[text1]" and "[text2]", do these two sentences have the same semantics?',
    "mnli":'Given the sentence "[text1]" determine if the following statement is entailed or contradicted or neutral: "[text2]"',
    "qnli":'Given the question "[text1]" determine if the following sentence contains the corresponding answer: "[text2]"',
    "rte":'Given the sentence "[text1]" determine if the following statement is entailed: "[text2]"',
    "wnli":'Given the sentence "[text1]" determine if the following statement is entailed: "[text2]"',
}

def  write_data(all_samples,save_path, sheet_name):
    if os.path.exists(save_path):
        bg = op.load_workbook(save_path)
        bg.create_sheet(sheet_name)
    else:
        bg = op.Workbook()
        bg.create_sheet(sheet_name)

    sheet = bg[sheet_name]            
    for c in range(1, len(list(all_samples[0].keys()))+1):
        key=list(all_samples[0].keys())[c-1]
        sheet.cell(row=1 , column=c,value=key)
        for r in range(2, len(all_samples)+2):	
            sheet.cell(row=r , column=c,value=all_samples[r-2][key])
    sheet.cell(row=1 , column=len(list(all_samples[0].keys()))+1,value="pred_ChatGPT")			

    bg.save(save_path)
    bg.close()
    return 

def get_data(args):
    k=args.num
    task=args.task
    results=args.model_pred
    data_path=task2data[task]
    template=task2prompt[task]
    save_path="{}/{}-dev.xlsx".format(args.save_path, task)
    prior_data_path="{}/{}-dev.xlsx".format(args.save_path, task)  # load the prior saved data, and re-use this data during getting the predictions made by the new model 

    if task == "cola":
        data_all=[]
        positive_data, negative_data=[], []
        ids=0
        with open(data_path) as d, open(results) as r:
            for line, pred in zip(d.readlines(), r.readlines()[1:]):
                data=line.strip().split("\t") 
                pred=pred.strip().split("\t")
                assert ids == int(pred[0])

                data_all.append({"id":ids, "sentence":data[3],"sentence4ChatGPT": template.replace("[text]", data[3]), "label (1 for yes, 0 for no)":data[1], "pred_deberta":pred[1]})
                if data[1]=="1":
                    positive_data.append({"id":ids, "sentence":data[3],"sentence4ChatGPT": template.replace("[text]", data[3]), "label (1 for yes, 0 for no)":data[1], "pred_deberta":pred[1]})
                else:
                    negative_data.append({"id":ids, "sentence":data[3],"sentence4ChatGPT": template.replace("[text]", data[3]), "label (1 for yes, 0 for no)":data[1], "pred_deberta":pred[1]})
                ids+=1

        if prior_data_path is not None:
            sheet_ids=[]
            bg = op.load_workbook(prior_data_path)
            sheet=bg["Sheet"]
            sheet_ids=[sheet.cell(i+2,1).value for i in range(2*k)]
            all_samples=[data_all[id] for id in sheet_ids]
        else:
            pos_samples=sample(positive_data, k)
            neg_samples=sample(negative_data, k)
            all_samples=pos_samples+neg_samples

        write_data(all_samples, save_path, "{}-{}".format(task, model))

    elif task == "sst2":
        data_all=[]
        positive_data, negative_data=[], []
        ids=0
        with open(data_path) as d, open(results) as r:
            for line, pred in zip(d.readlines()[1:], r.readlines()[1:]):
                data=line.strip().split("\t") 
                pred=pred.strip().split("\t")
                assert ids == int(pred[0])

                data_all.append({"id":ids, "sentence":data[0],"sentence4ChatGPT": template.replace("[text]", data[0]), "label (1 for yes, 0 for no)":data[1], "pred_deberta":pred[1]})
                if data[1]=="1":
                    positive_data.append({"id":ids, "sentence":data[0],"sentence4ChatGPT": template.replace("[text]", data[0]), "label (1 for yes, 0 for no)":data[1], "pred_deberta":pred[1]})
                else:
                    negative_data.append({"id":ids, "sentence":data[0],"sentence4ChatGPT": template.replace("[text]", data[0]), "label (1 for yes, 0 for no)":data[1], "pred_deberta":pred[1]})
                ids+=1

        if prior_data_path is not None:
            sheet_ids=[]
            bg = op.load_workbook(prior_data_path)
            sheet=bg["Sheet"]
            sheet_ids=[sheet.cell(i+2,1).value for i in range(2*k)]
            all_samples=[data_all[id] for id in sheet_ids]
        else:
            pos_samples=sample(positive_data, k)
            neg_samples=sample(negative_data, k)
            all_samples=pos_samples+neg_samples

        write_data(all_samples, save_path, "{}-{}".format(task, model))

    elif task == "rte":
        data_all=[]
        positive_data, negative_data=[], []
        ids=0
        with open(data_path) as d, open(results) as r:
            for line, pred in zip(d.readlines()[1:], r.readlines()[1:]):
                data=line.strip().split("\t") 
                pred=pred.strip().split("\t")
                assert ids == int(pred[0])

                data_all.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label":data[-1], "pred_bert-base (0 for entailment, 0 for un_entailment)":pred[1]})
                if data[-1]=="entailment":
                    positive_data.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label":data[-1], "pred_bert-base (0 for entailment, 0 for un_entailment)":pred[1]})
                else:
                    negative_data.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label":data[-1], "pred_bert-base (0 for entailment, 0 for un_entailment)":pred[1]})
                ids+=1

        if prior_data_path is not None:
            sheet_ids=[]
            bg = op.load_workbook(prior_data_path)
            sheet=bg["Sheet"]
            sheet_ids=[sheet.cell(i+2,1).value for i in range(2*k)]
            all_samples=[data_all[id] for id in sheet_ids]
        else:
            pos_samples=sample(positive_data, k)
            neg_samples=sample(negative_data, k)
            all_samples=pos_samples+neg_samples

        write_data(all_samples, save_path, "{}-{}".format(task, model))

    elif task == "mrpc":
        data_all=[]
        positive_data, negative_data=[], []
        ids=0
        with open(data_path) as d, open(results) as r:
            for line, pred in zip(d.readlines()[1:], r.readlines()[1:]):
                data=line.strip().split("\t") 
                pred=pred.strip().split("\t")
                assert ids == int(pred[0])

                data_all.append({"id":ids, "sentence1":data[3],"sentence2":data[4],"sentence4ChatGPT": template.replace("[text1]", data[3]).replace("[text2]", data[4]), "label (1 for yes, 0 for no)":data[0], "pred_bert-base":pred[1]})
                if data[0]=="1":
                    positive_data.append({"id":ids, "sentence1":data[3],"sentence2":data[4],"sentence4ChatGPT": template.replace("[text1]", data[3]).replace("[text2]", data[4]), "label (1 for yes, 0 for no)":data[0], "pred_bert-base":pred[1]})
                else:
                    negative_data.append({"id":ids, "sentence1":data[3],"sentence2":data[4],"sentence4ChatGPT": template.replace("[text1]", data[3]).replace("[text2]", data[4]), "label (1 for yes, 0 for no)":data[0], "pred_bert-base":pred[1]})
                ids+=1

        if prior_data_path is not None:
            sheet_ids=[]
            bg = op.load_workbook(prior_data_path)
            sheet=bg["Sheet"]
            sheet_ids=[sheet.cell(i+2,1).value for i in range(2*k)]
            all_samples=[data_all[id] for id in sheet_ids]
        else:
            pos_samples=sample(positive_data, k)
            neg_samples=sample(negative_data, k)
            all_samples=pos_samples+neg_samples

        write_data(all_samples, save_path, "{}-{}".format(task, model))

    elif task =="mnli":
        for i in range(2):
            data_all=[]
            positive_data, negative_data, neutral_data=[], [], []
            ids=0
            with open(data_path[i]) as d, open(results[i]) as r:
                for line, pred in zip(d.readlines()[1:], r.readlines()[1:]):
                    data=line.strip().split("\t") 
                    pred=pred.strip().split("\t")
                    assert ids == int(pred[0])

                    data_all.append({"id":ids, "sentence1":data[-8],"sentence2":data[-7],"sentence4ChatGPT": template.replace("[text1]", data[-8]).replace("[text2]", data[-7]), "label (2 for entailment, 1 for neutral, 0 for contradtion)":data[-1], "pred_bert-base":pred[1]})
                    if data[-1]=="neutral":
                        neutral_data.append({"id":ids, "sentence1":data[-8],"sentence2":data[-7],"sentence4ChatGPT": template.replace("[text1]", data[-8]).replace("[text2]", data[-7]), "label (2 for entailment, 1 for neutral, 0 for contradtion)":data[-1], "pred_bert-base":pred[1]})
                    elif data[-1]=="entailment":
                        positive_data.append({"id":ids, "sentence1":data[-8],"sentence2":data[-7],"sentence4ChatGPT": template.replace("[text1]", data[-8]).replace("[text2]", data[-7]), "label (2 for entailment, 1 for neutral, 0 for contradtion)":data[-1], "pred_bert-base":pred[1]})
                    else:
                        negative_data.append({"id":ids, "sentence1":data[-8],"sentence2":data[-7],"sentence4ChatGPT": template.replace("[text1]", data[-8]).replace("[text2]", data[-7]), "label (2 for entailment, 1 for neutral, 0 for contradtion)":data[-1], "pred_bert-base":pred[1]})
                    ids+=1

            if prior_data_path is not None:
                sheet_ids=[]
                bg = op.load_workbook(prior_data_path)
                sheet=bg["Sheet-m" if i ==0 else "Sheet-mm"]
                sheet_ids=[sheet.cell(i+2,1).value for i in range(3*k)]
                all_samples=[data_all[id] for id in sheet_ids]
            else:
                pos_samples=sample(positive_data, k)
                neg_samples=sample(negative_data, k)
                neu_samples=sample(neutral_data, k)
                all_samples=pos_samples+neg_samples+neu_samples

            if i == 1:
                write_data(all_samples, save_path, "{}-m-{}".format(task, model))
            else:
                write_data(all_samples, save_path, "{}-mm-{}".format(task, model))

    elif task == "stsb":
        data_all=[]
        positive_data, negative_data=[], []
        ids=0
        with open(data_path) as d, open(results) as r:
            for line, pred in zip(d.readlines()[1:], r.readlines()[1:]):
                data=line.strip().split("\t") 
                pred=pred.strip().split("\t")
                assert ids == int(pred[0]), "{},{}".format(ids, pred[0])

                data_all.append({"id":ids, "sentence1":data[-3],"sentence2":data[-2],"sentence4ChatGPT": template.replace("[text1]", data[-3]).replace("[text2]", data[-2]), "label (2 for entailment, 1 for neutral, 0 for contradtion)":data[-1], "pred_bert-base":pred[1]})
                ids+=1

        if prior_data_path is not None:
            sheet_ids=[]
            bg = op.load_workbook(prior_data_path)
            sheet=bg["Sheet"]
            sheet_ids=[sheet.cell(i+2,1).value for i in range(2*k)]
            all_samples=[data_all[id] for id in sheet_ids]
        else:
            all_samples=sample(data_all, 2*k)

        write_data(all_samples, save_path, "{}-{}".format(task, model))

    elif task =="qnli":
        data_all=[]
        positive_data, negative_data=[], []
        ids=0
        with open(data_path) as d, open(results) as r:
            for line, pred in zip(d.readlines()[1:], r.readlines()[1:]):
                data=line.strip().split("\t") 
                pred=pred.strip().split("\t")
                assert ids == int(pred[0])

                data_all.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label (1 for yes, 0 for no)":data[3], "pred_bert-base":pred[1]})
                if data[3]=="entailment":
                    positive_data.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label (1 for yes, 0 for no)":data[3], "pred_bert-base":pred[1]})
                else:
                    negative_data.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label (1 for yes, 0 for no)":data[3], "pred_bert-base":pred[1]})
                ids+=1

        if prior_data_path is not None:
            sheet_ids=[]
            bg = op.load_workbook(prior_data_path)
            sheet=bg["Sheet"]
            sheet_ids=[sheet.cell(i+2,1).value for i in range(2*k)]
            all_samples=[data_all[id] for id in sheet_ids]
        else:
            pos_samples=sample(positive_data, k)
            neg_samples=sample(negative_data, k)
            all_samples=pos_samples+neg_samples

        write_data(all_samples, save_path, "{}-{}".format(task, model))

    elif task =="qqp":
        data_all=[]
        positive_data, negative_data=[], []
        ids=0
        with open(data_path) as d, open(results) as r:
            for line, pred in zip(d.readlines()[1:], r.readlines()[1:]):
                data=line.strip().split("\t") 
                pred=pred.strip().split("\t")
                assert ids == int(pred[0])

                data_all.append({"id":ids, "sentence1":data[-3],"sentence2":data[-2],"sentence4ChatGPT": template.replace("[text1]", data[-3]).replace("[text2]", data[-2]), "label (1 for yes, 0 for no)":data[-1], "pred_bert-base":pred[1]})
                if data[-1]=="1":
                    positive_data.append({"id":ids, "sentence1":data[-3],"sentence2":data[-2],"sentence4ChatGPT": template.replace("[text1]", data[-3]).replace("[text2]", data[-2]), "label (1 for yes, 0 for no)":data[-1], "pred_bert-base":pred[1]})
                else:
                    negative_data.append({"id":ids, "sentence1":data[-3],"sentence2":data[-2],"sentence4ChatGPT": template.replace("[text1]", data[-3]).replace("[text2]", data[-2]), "label (1 for yes, 0 for no)":data[-1], "pred_bert-base":pred[1]})
                ids+=1

        if prior_data_path is not None:
            sheet_ids=[]
            bg = op.load_workbook(prior_data_path)
            sheet=bg["Sheet"]
            sheet_ids=[sheet.cell(i+2,1).value for i in range(2*k)]
            all_samples=[data_all[id] for id in sheet_ids]
        else:
            pos_samples=sample(positive_data, k)
            neg_samples=sample(negative_data, k)
            all_samples=pos_samples+neg_samples

        write_data(all_samples, save_path, "{}-{}".format(task, model))

    elif task=="wnli":
        data_all=[]
        positive_data, negative_data=[], []
        ids=0
        with open(data_path) as d, open(results) as r:
            for line, pred in zip(d.readlines()[1:], r.readlines()[1:]):
                data=line.strip().split("\t") 
                pred=pred.strip().split("\t")
                assert ids == int(pred[0])

                data_all.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label (1 for yes, 0 for no)":data[3], "pred_bert-base":pred[1]})
                if data[3]=="entailment":
                    positive_data.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label (1 for yes, 0 for no)":data[3], "pred_bert-base":pred[1]})
                else:
                    negative_data.append({"id":ids, "sentence1":data[1],"sentence2":data[2],"sentence4ChatGPT": template.replace("[text1]", data[1]).replace("[text2]", data[2]), "label (1 for yes, 0 for no)":data[3], "pred_bert-base":pred[1]})
                ids+=1

        if prior_data_path is not None:
            sheet_ids=[]
            bg = op.load_workbook(prior_data_path)
            sheet=bg["Sheet"]
            sheet_ids=[sheet.cell(i+2,1).value for i in range(2*k)]
            all_samples=[data_all[id] for id in sheet_ids]
        else:
            pos_samples=sample(positive_data, k)
            neg_samples=sample(negative_data, k)
            all_samples=pos_samples+neg_samples

        write_data(all_samples, save_path, "{}-{}".format(task, model))

    else:
        print("Error! Please input the correct task name, from the following choices:{}".format(list(task2data.keys())))

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--num', type=int, default=25, help='max num')
    parser.add_argument('--task', type=str, required=True, help='task name')
    parser.add_argument('--model_pred', type=str, required=True, help='the path of model prediction, where each line contains the id and prediction.')
    parser.add_argument('--save_path', type=str, default="./data", help='save path')
    args = parser.parse_args()

    assert args.task in list(task2data.keys()), "Error! Please input the correct task name, from the following choices:{}".format(list(task2data.keys()))
    get_data(args)
